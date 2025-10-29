#!/usr/bin/env python3
"""
csv_editor_streaming_final.py

Terminal streaming CSV/XLSX editor (curses) with:
- Low-memory line-by-line streaming for structural edits
- Curses UI with fixed 50-row viewport
- Multi-cell selection: ';' to start, arrows to expand, Enter confirm (copy), Esc cancel
- Copy/Paste (overwrite) using internal clipboard
- Insert/Delete rows & columns (deterministic behavior)
- Search with '/', navigate results with 'n'/'N'
- Atomic save with .bak; final save in original format (.csv or .xlsx)
- Enforces minimum 1 row x 1 column (prevents destructive deletes)
- Structural ops wait until the stream worker applies them (prevent phantom UI)
"""

import os
import sys
import csv
import time
import shutil
import threading
import multiprocessing
import curses
from concurrent.futures import ProcessPoolExecutor
from dataclasses import dataclass
from typing import List, Tuple, Optional, Dict, Set

# Optional libs (used where available)
try:
    import openpyxl
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

# ---------- small helpers ----------
def atomic_replace(src: str, dst: str):
    """Replace dst with src atomically, keep a .bak copy of dst if exists."""
    bak = dst + ".bak"
    try:
        if os.path.exists(dst):
            shutil.copy2(dst, bak)
        os.replace(src, dst)
        if os.path.exists(bak):
            os.remove(bak)
    except Exception:
        # restore if failure
        if os.path.exists(bak):
            os.replace(bak, dst)
        raise

def ensure_file(path: str):
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)
    if not os.path.exists(path):
        open(path, 'w', encoding='utf-8').close()

def is_xlsx(path: str) -> bool:
    return path.lower().endswith('.xlsx')

def is_csv(path: str) -> bool:
    return path.lower().endswith('.csv')

@dataclass
class CellModification:
    row: int
    col: int
    old_value: str
    new_value: str
    ts: float

# ---------- Streaming File Manager ----------
class StreamingFileManager:
    """Low-memory line-by-line streaming operations on a base CSV file (temp file)."""
    def __init__(self, base_path: str):
        self.orig_path = base_path
        self.temp_file = base_path + '.stream_tmp.csv'
        ensure_file(self.orig_path)
        ensure_file(self.temp_file)
        try:
            if os.path.exists(self.orig_path) and os.path.getsize(self.temp_file) == 0:
                shutil.copy2(self.orig_path, self.temp_file)
        except Exception:
            pass

        self._seq_lock = threading.Lock()
        self._seq_cond = threading.Condition()
        self.change_seq = 0        # next seq to assign
        self.applied_seq = 0       # next seq expected to be applied
        self.pending: Dict[int, Tuple[str, object]] = {}
        self.file_lock = threading.Lock()
        self.streaming_active = True

        self.max_workers = min(4, multiprocessing.cpu_count())
        self.pool = ProcessPoolExecutor(max_workers=self.max_workers)

        self.worker = threading.Thread(target=self._worker_loop, daemon=True)
        self.worker.start()

    def _worker_loop(self):
        while self.streaming_active:
            with self._seq_cond:
                while self.applied_seq not in self.pending and self.streaming_active:
                    self._seq_cond.wait(timeout=0.1)
                if not self.streaming_active:
                    break
                seq = self.applied_seq
                entry = self.pending.pop(seq, None)
            if entry is None:
                continue
            kind, payload = entry
            try:
                with self.file_lock:
                    if kind == 'cell':
                        self._apply_cell(payload)
                    elif kind == 'row':
                        op, idx = payload
                        if op == 'insert':
                            self._insert_row_stream(idx)
                        else:
                            self._delete_row_stream(idx)
                    elif kind == 'col':
                        op, idx = payload
                        if op == 'insert':
                            self._insert_col_stream(idx)
                        else:
                            self._delete_col_stream(idx)
                # mark applied and notify waiters
                with self._seq_cond:
                    self.applied_seq += 1
                    self._seq_cond.notify_all()
            except Exception as e:
                # log and advance to avoid deadlock
                print(f"stream worker failed seq {seq}: {e}", file=sys.stderr)
                with self._seq_cond:
                    self.applied_seq += 1
                    self._seq_cond.notify_all()

    # Queueing API (returns sequence id)
    def queue_cell_change(self, row:int, col:int, old_value:str, new_value:str) -> int:
        mod = CellModification(row, col, old_value, new_value, time.time())
        with self._seq_lock:
            seq = self.change_seq
            self.change_seq += 1
            self.pending[seq] = ('cell', mod)
        with self._seq_cond:
            self._seq_cond.notify_all()
        return seq

    def queue_row_op(self, op:str, index:int) -> int:
        if op not in ('insert','delete'):
            raise ValueError("invalid row op")
        with self._seq_lock:
            seq = self.change_seq
            self.change_seq += 1
            self.pending[seq] = ('row', (op, index))
        with self._seq_cond:
            self._seq_cond.notify_all()
        return seq

    def queue_col_op(self, op:str, index:int) -> int:
        if op not in ('insert','delete'):
            raise ValueError("invalid col op")
        with self._seq_lock:
            seq = self.change_seq
            self.change_seq += 1
            self.pending[seq] = ('col', (op, index))
        with self._seq_cond:
            self._seq_cond.notify_all()
        return seq

    def wait_for_seq(self, seq:int, timeout: Optional[float] = 10.0) -> bool:
        """Block until the given seq has been applied (or timeout)."""
        deadline = time.time() + (timeout if timeout is not None else 10.0)
        with self._seq_cond:
            while self.applied_seq <= seq and time.time() < deadline:
                remaining = deadline - time.time()
                self._seq_cond.wait(timeout=remaining if remaining>0 else 0.0)
            return self.applied_seq > seq

    # ----------------- streaming implementations -----------------
    def _apply_cell(self, mod: CellModification):
        tmp = self.temp_file + f".cell_{int(time.time()*1000)}.tmp"
        try:
            with open(self.temp_file, 'r', encoding='utf-8', newline='') as src, \
                 open(tmp, 'w', encoding='utf-8', newline='') as dst:
                r = csv.reader(src)
                w = csv.writer(dst)
                for i, row in enumerate(r):
                    if i == mod.row:
                        while len(row) <= mod.col:
                            row.append('')
                        row[mod.col] = mod.new_value
                    w.writerow(row)
            atomic_replace(tmp, self.temp_file)
        finally:
            if os.path.exists(tmp):
                try: os.remove(tmp)
                except: pass

    def _insert_row_stream(self, insert_index:int):
        tmp = self.temp_file + f".rowins_{int(time.time()*1000)}.tmp"
        try:
            rows = []
            maxcols = 0
            with open(self.temp_file, 'r', encoding='utf-8', newline='') as f:
                r = csv.reader(f)
                for row in r:
                    rows.append(row)
                    if len(row) > maxcols:
                        maxcols = len(row)
            if maxcols == 0:
                maxcols = 1
            idx = max(0, min(insert_index, len(rows)))
            with open(tmp, 'w', encoding='utf-8', newline='') as dst:
                w = csv.writer(dst)
                for i, row in enumerate(rows):
                    if i == idx:
                        w.writerow([''] * maxcols)
                    w.writerow(row)
                if idx == len(rows):
                    w.writerow([''] * maxcols)
            atomic_replace(tmp, self.temp_file)
        finally:
            if os.path.exists(tmp):
                try: os.remove(tmp)
                except: pass

    def _delete_row_stream(self, delete_index:int):
        tmp = self.temp_file + f".rowdel_{int(time.time()*1000)}.tmp"
        try:
            total = 0
            with open(self.temp_file, 'r', encoding='utf-8', newline='') as f:
                for _ in csv.reader(f):
                    total += 1
            if total <= 1:
                # blocked by rule
                return False
            with open(self.temp_file, 'r', encoding='utf-8', newline='') as src, \
                 open(tmp, 'w', encoding='utf-8', newline='') as dst:
                r = csv.reader(src)
                w = csv.writer(dst)
                for i, row in enumerate(r):
                    if i == delete_index:
                        continue
                    w.writerow(row)
            atomic_replace(tmp, self.temp_file)
            return True
        finally:
            if os.path.exists(tmp):
                try: os.remove(tmp)
                except: pass

    def _insert_col_stream(self, insert_index:int):
        tmp = self.temp_file + f".colins_{int(time.time()*1000)}.tmp"
        try:
            maxcols = 0
            with open(self.temp_file, 'r', encoding='utf-8', newline='') as f:
                for row in csv.reader(f):
                    if len(row) > maxcols:
                        maxcols = len(row)
            if maxcols == 0:
                maxcols = 1
            with open(self.temp_file, 'r', encoding='utf-8', newline='') as src, \
                 open(tmp, 'w', encoding='utf-8', newline='') as dst:
                r = csv.reader(src)
                w = csv.writer(dst)
                for row in r:
                    if len(row) < maxcols:
                        row += [''] * (maxcols - len(row))
                    pos = max(0, min(insert_index + 1, len(row)))
                    row.insert(pos, '')
                    w.writerow(row)
            atomic_replace(tmp, self.temp_file)
        finally:
            if os.path.exists(tmp):
                try: os.remove(tmp)
                except: pass

    def _delete_col_stream(self, delete_index:int):
        tmp = self.temp_file + f".coldel_{int(time.time()*1000)}.tmp"
        try:
            maxcols = 0
            with open(self.temp_file, 'r', encoding='utf-8', newline='') as f:
                for row in csv.reader(f):
                    if len(row) > maxcols:
                        maxcols = len(row)
            if maxcols <= 1:
                return False
            with open(self.temp_file, 'r', encoding='utf-8', newline='') as src, \
                 open(tmp, 'w', encoding='utf-8', newline='') as dst:
                r = csv.reader(src)
                w = csv.writer(dst)
                for row in r:
                    if 0 <= delete_index < len(row):
                        del row[delete_index]
                    w.writerow(row)
            atomic_replace(tmp, self.temp_file)
            return True
        finally:
            if os.path.exists(tmp):
                try: os.remove(tmp)
                except: pass

    def finalize_and_write(self, out_path: str, format_hint: str = 'csv', timeout: Optional[float] = None) -> bool:
        start = time.time()
        while True:
            with self._seq_cond:
                pending_empty = len(self.pending) == 0
                done = self.applied_seq >= self.change_seq
            if pending_empty and done:
                break
            if timeout is not None and (time.time() - start) > timeout:
                raise TimeoutError("Timeout waiting for streaming apply")
            time.sleep(0.02)

        maxcols = 0
        with open(self.temp_file, 'r', encoding='utf-8', newline='') as f:
            for row in csv.reader(f):
                if len(row) > maxcols:
                    maxcols = len(row)
        if maxcols == 0:
            maxcols = 1

        final_tmp = self.temp_file + '.finaltmp'
        with open(self.temp_file, 'r', encoding='utf-8', newline='') as src, \
             open(final_tmp, 'w', encoding='utf-8', newline='') as dst:
            r = csv.reader(src)
            w = csv.writer(dst)
            for row in r:
                if len(row) < maxcols:
                    row += [''] * (maxcols - len(row))
                elif len(row) > maxcols:
                    row = row[:maxcols]
                w.writerow(row)

        try:
            try:
                if os.path.exists(out_path):
                    shutil.copy2(out_path, out_path + '.bak')
            except Exception:
                pass

            if format_hint == 'csv' or not HAS_OPENPYXL:
                atomic_replace(final_tmp, out_path)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                with open(final_tmp, 'r', encoding='utf-8', newline='') as f:
                    for rrow in csv.reader(f):
                        ws.append([ '' if v is None else v for v in rrow ])
                wb.save(out_path)
                try:
                    os.remove(final_tmp)
                except Exception:
                    pass
            return True
        except Exception as e:
            try:
                if os.path.exists(out_path + '.bak'):
                    atomic_replace(out_path + '.bak', out_path)
            except Exception:
                pass
            print(f"[finalize] error: {e}", file=sys.stderr)
            return False

    def cleanup(self):
        """Gracefully stop streaming workers and remove temporary files."""
        # --- Stop background threads and workers ---
        self.streaming_active = False
        try:
            with self._seq_cond:
                self._seq_cond.notify_all()
        except Exception:
            pass

        try:
            if hasattr(self, "worker") and self.worker.is_alive():
                self.worker.join(timeout=1.0)
        except Exception:
            pass

        try:
            if hasattr(self, "pool"):
                self.pool.shutdown(wait=False)
        except Exception:
            pass

        # --- Clean up temp/backup/arrow files ---
        try:
            base_dir = os.path.dirname(self.orig_path) or "."
            base_name = os.path.splitext(os.path.basename(self.orig_path))[0]

            candidates = set()
            for suffix in [".stream_tmp.csv", ".bak", ".tmp", ".arrow"]:
                path = os.path.join(base_dir, base_name + suffix)
                if os.path.exists(path):
                    candidates.add(path)

            # Also remove residuals matching base prefix
            for f in os.listdir(base_dir):
                if f.startswith(base_name) and f.endswith((".stream_tmp.csv", ".tmp", ".bak", ".arrow")):
                    candidates.add(os.path.join(base_dir, f))

            for path in candidates:
                try:
                    os.remove(path)
                except Exception:
                    pass
        except Exception:
            pass

# ---------- Editor (UI-facing) ----------
class Editor:
    VIEW_ROWS = 50
    VIEW_COLS = 8

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.orig_format = 'xlsx' if is_xlsx(filepath) else 'csv'
        self.fileman = StreamingFileManager(filepath)
        if self.orig_format == 'xlsx':
            if not HAS_OPENPYXL:
                raise RuntimeError("openpyxl required for xlsx support")
            self._xlsx_to_temp_csv()

        self.current_row = 0
        self.current_col = 0
        self.offset_row = 0
        self.offset_col = 0
        self.visible_rows: List[List[str]] = []
        self.total_rows = 0
        self.max_cols = 1

        self.modifications: Dict[Tuple[int,int], str] = {}
        self.deleted_cells: Set[Tuple[int,int]] = set()
        self.modified = False
        self.clipboard: List[List[str]] = []

        self.selection_active = False
        self.sel_start: Optional[Tuple[int,int]] = None
        self.sel_end: Optional[Tuple[int,int]] = None

        self.search_hits: List[Tuple[int,int]] = []
        self.search_index = -1

        self._recompute_metadata()
        self.reload_visible()
        self.status = "Ready"

    def _xlsx_to_temp_csv(self):
        tmp = self.fileman.temp_file
        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True)
            ws = wb.active
            with open(tmp, 'w', encoding='utf-8', newline='') as f:
                w = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    w.writerow([ '' if v is None else str(v) for v in row ])
            try:
                shutil.copy2(tmp, self.fileman.temp_file)
            except Exception:
                pass
        except Exception as e:
            raise RuntimeError(f"xlsx->csv conversion failed: {e}")

    def _recompute_metadata(self):
        rows = 0
        maxc = 0
        with open(self.fileman.temp_file, 'r', encoding='utf-8', newline='') as f:
            r = csv.reader(f)
            for row in r:
                rows += 1
                if len(row) > maxc:
                    maxc = len(row)
        if rows == 0:
            rows = 1
        if maxc == 0:
            maxc = 1
        self.total_rows = rows
        self.max_cols = maxc
        self.current_row = max(0, min(self.current_row, self.total_rows - 1))
        self.current_col = max(0, min(self.current_col, self.max_cols - 1))

    def reload_visible(self):
        self.visible_rows = []
        start = self.offset_row
        end = start + self.VIEW_ROWS
        with open(self.fileman.temp_file, 'r', encoding='utf-8', newline='') as f:
            r = csv.reader(f)
            for i, row in enumerate(r):
                if i < start:
                    continue
                if i >= end:
                    break
                self.visible_rows.append(row)
        for i, row in enumerate(self.visible_rows):
            if len(row) < self.max_cols:
                row += [''] * (self.max_cols - len(row))
            self.visible_rows[i] = row

    def ensure_visible(self):
        """Ensure the current cell is visible in the viewport."""
        # Vertical adjustment
        if self.current_row < self.offset_row:
            self.offset_row = self.current_row
            self.reload_visible()
        elif self.current_row >= self.offset_row + self.VIEW_ROWS:
            self.offset_row = self.current_row - self.VIEW_ROWS + 1
            self.reload_visible()
        # Horizontal adjustment
        if self.current_col < self.offset_col:
            self.offset_col = self.current_col
            self.reload_visible()
        elif self.current_col >= self.offset_col + self.VIEW_COLS:
            self.offset_col = self.current_col - self.VIEW_COLS + 1
            self.reload_visible()

    def get_cell(self, row:int, col:int) -> str:
        if (row,col) in self.modifications:
            return self.modifications[(row,col)]
        if (row,col) in self.deleted_cells:
            return ''
        if self.offset_row <= row < self.offset_row + len(self.visible_rows):
            r = row - self.offset_row
            if col < len(self.visible_rows[r]):
                return self.visible_rows[r][col]
            return ''
        with open(self.fileman.temp_file, 'r', encoding='utf-8', newline='') as f:
            r = csv.reader(f)
            for i, rr in enumerate(r):
                if i == row:
                    return rr[col] if col < len(rr) else ''
        return ''

    def set_cell(self, row:int, col:int, value: str):
        old = self.get_cell(row,col)
        self.modifications[(row,col)] = value
        self.deleted_cells.discard((row,col))
        self.modified = True
        self.fileman.queue_cell_change(row, col, old, value)
        if self.offset_row <= row < self.offset_row + len(self.visible_rows):
            r = row - self.offset_row
            while len(self.visible_rows[r]) <= col:
                self.visible_rows[r].append('')
            self.visible_rows[r][col] = value

    def clear_cell(self, row:int, col:int):
        old = self.get_cell(row,col)
        self.modifications.pop((row,col), None)
        self.deleted_cells.add((row,col))
        self.modified = True
        self.fileman.queue_cell_change(row, col, old, '')
        if self.offset_row <= row < self.offset_row + len(self.visible_rows):
            r = row - self.offset_row
            if col < len(self.visible_rows[r]):
                self.visible_rows[r][col] = ''

    # Structural ops: queue op, wait for seq to apply, then refresh metadata to avoid phantom display.
    def insert_row_below(self):
        idx = self.current_row + 1
        seq = self.fileman.queue_row_op('insert', idx)
        self.fileman.wait_for_seq(seq, timeout=10.0)
        self._refresh_after_edit()
        self.current_row = idx
        if self.current_row >= self.offset_row + self.VIEW_ROWS:
            self.offset_row = self.current_row - self.VIEW_ROWS + 1
        self.reload_visible()
        self.status = f"Inserted row at {idx+1}"
        self.modified = True

    def insert_row_above(self):
        idx = self.current_row
        seq = self.fileman.queue_row_op('insert', idx)
        self.fileman.wait_for_seq(seq, timeout=10.0)
        self._refresh_after_edit()
        if self.current_row < self.offset_row:
            self.offset_row = self.current_row
        self.reload_visible()
        self.status = f"Inserted row at {idx+1}"
        self.modified = True

    def delete_current_row(self):
        idx = self.current_row
        total_before = self.total_rows
        seq = self.fileman.queue_row_op('delete', idx)
        self.fileman.wait_for_seq(seq, timeout=10.0)
        self._refresh_after_edit()
        if self.total_rows == total_before and total_before <= 1:
            self.status = "Cannot delete — minimum 1 row required"
            return
        if self.current_row > 0:
            self.current_row -= 1
        if self.current_row < self.offset_row:
            self.offset_row = max(0, self.current_row)
        self.reload_visible()
        self.status = f"Deleted row {idx+1}"
        self.modified = True

    def insert_col_right(self):
        idx = self.current_col
        seq = self.fileman.queue_col_op('insert', idx)
        self.fileman.wait_for_seq(seq, timeout=10.0)
        self._refresh_after_edit()
        self.current_col = idx + 1
        if self.current_col >= self.offset_col + self.VIEW_COLS:
            self.offset_col = self.current_col - self.VIEW_COLS + 1
        self.reload_visible()
        self.status = f"Inserted column at {self.current_col+1}"
        self.modified = True

    def insert_col_left(self):
        idx = self.current_col
        seq = self.fileman.queue_col_op('insert', idx - 1)
        self.fileman.wait_for_seq(seq, timeout=10.0)
        self._refresh_after_edit()
        if self.current_col < self.offset_col:
            self.offset_col = self.current_col
        self.reload_visible()
        self.status = f"Inserted column at {self.current_col+1}"
        self.modified = True

    def delete_current_col(self):
        idx = self.current_col
        if self.max_cols <= 1:
            self.status = "Cannot delete — minimum 1 column required"
            return
        seq = self.fileman.queue_col_op('delete', idx)
        self.fileman.wait_for_seq(seq, timeout=10.0)
        self._refresh_after_edit()
        if self.current_col > 0:
            self.current_col -= 1
        if self.current_col < self.offset_col:
            self.offset_col = max(0, self.current_col)
        self.reload_visible()
        self.status = f"Deleted column {idx+1}"
        self.modified = True

    # selection & clipboard
    def start_selection(self):
        self.selection_active = True
        self.sel_start = (self.current_row, self.current_col)
        self.sel_end = (self.current_row, self.current_col)
        self.status = "Selection started"

    def expand_selection_to_cursor(self):
        if self.selection_active:
            self.sel_end = (self.current_row, self.current_col)

    def cancel_selection(self):
        self.selection_active = False
        self.sel_start = None
        self.sel_end = None
        self.status = "Selection cancelled"

    def confirm_selection_copy(self):
        if not self.selection_active or self.sel_start is None or self.sel_end is None:
            self.status = "No selection"
            return
        top = min(self.sel_start[0], self.sel_end[0])
        bottom = max(self.sel_start[0], self.sel_end[0])
        left = min(self.sel_start[1], self.sel_end[1])
        right = max(self.sel_start[1], self.sel_end[1])
        out = []
        with open(self.fileman.temp_file, 'r', encoding='utf-8', newline='') as f:
            r = csv.reader(f)
            for i, row in enumerate(r):
                if i < top: continue
                if i > bottom: break
                if len(row) <= right:
                    row = row + [''] * (right + 1 - len(row))
                out.append([ row[j] if j < len(row) else '' for j in range(left, right + 1) ])
        self.clipboard = out
        self.selection_active = False
        self.sel_start = self.sel_end = None
        self.status = f"Copied {len(out)}×{(right-left+1)} block"

    def copy_single_or_selection(self):
        if self.selection_active:
            self.confirm_selection_copy()
            return
        self.clipboard = [[ self.get_cell(self.current_row, self.current_col) ]]
        self.status = "Copied current cell"

    def paste_overwrite(self):
        if not self.clipboard:
            return []
        pasted_cells = []
        start_r, start_c = self.current_row, self.current_col
        for i, rowvals in enumerate(self.clipboard):
            for j, val in enumerate(rowvals):
                r = start_r + i
                c = start_c + j
                self.set_cell(r, c, val)
                pasted_cells.append((r, c, val))
        self.modified = True
        return pasted_cells

    # search
    def search(self, term: str):
        self.search_hits = []
        if not term:
            self.status = "Empty search"
            return
        low = term.lower()
        try:
            total = self.total_rows
            workers = max(1, self.fileman.max_workers)
            chunk = max(1000, total // (workers * 4) or 1000)
        except Exception:
            chunk = 5000
        ranges = []
        s = 0
        while s < self.total_rows:
            e = min(self.total_rows, s + chunk)
            ranges.append((s,e))
            s = e
        futures = [ self.fileman.pool.submit(_search_chunk, self.fileman.temp_file, st, ed, low) for st,ed in ranges ]
        hits = []
        for fut in futures:
            try:
                res = fut.result(timeout=30)
                hits.extend(res)
            except Exception:
                pass
        hits.sort()
        self.search_hits = hits
        if hits:
            self.search_index = 0
            self.current_row, self.current_col = hits[0]
            if not (self.offset_row <= self.current_row < self.offset_row + self.VIEW_ROWS):
                self.offset_row = max(0, self.current_row - self.VIEW_ROWS//2)
            if not (self.offset_col <= self.current_col < self.offset_col + self.VIEW_COLS):
                self.offset_col = max(0, self.current_col - self.VIEW_COLS//2)
            self.reload_visible()
            self.status = f"Found {len(hits)} results"
        else:
            self.status = "No matches"

    def search_next(self):
        if not self.search_hits:
            self.status = "No search results"
            return

        # wrap around instead of stopping
        self.search_index = (self.search_index + 1) % len(self.search_hits)
        self.current_row, self.current_col = self.search_hits[self.search_index]

        # vertical visibility
        if not (self.offset_row <= self.current_row < self.offset_row + self.VIEW_ROWS):
            self.offset_row = max(0, self.current_row - self.VIEW_ROWS // 2)

        # horizontal visibility
        if not (self.offset_col <= self.current_col < self.offset_col + self.VIEW_COLS):
            self.offset_col = max(0, self.current_col - self.VIEW_COLS // 2)

        self.reload_visible()
        self.status = f"Search {self.search_index + 1}/{len(self.search_hits)}"


    def search_prev(self):
        if not self.search_hits:
            self.status = "No search results"
            return

        # wrap backward
        self.search_index = (self.search_index - 1) % len(self.search_hits)
        self.current_row, self.current_col = self.search_hits[self.search_index]

        # vertical visibility
        if not (self.offset_row <= self.current_row < self.offset_row + self.VIEW_ROWS):
            self.offset_row = max(0, self.current_row - self.VIEW_ROWS // 2)

        # horizontal visibility
        if not (self.offset_col <= self.current_col < self.offset_col + self.VIEW_COLS):
            self.offset_col = max(0, self.current_col - self.VIEW_COLS // 2)

        self.reload_visible()
        self.status = f"Search {self.search_index + 1}/{len(self.search_hits)}"

    def goto_row(self, row_number: int):
        """Jump directly to a specific row number (1-based)."""
        if row_number < 1:
            row_number = 1

        max_rows = self.total_rows
        if row_number > max_rows:
            row_number = max_rows

        self.current_row = row_number - 1
        # Adjust visible window
        if not (self.offset_row <= self.current_row < self.offset_row + self.VIEW_ROWS):
            self.offset_row = max(0, self.current_row - self.VIEW_ROWS // 2)
        self.reload_visible()
        self.status = f"Jumped to row {row_number}"

    def _refresh_after_edit(self):
        """Recompute metadata and keep UI consistent after structural edits."""
        self._recompute_metadata()
        self.current_row = max(0, min(self.current_row, self.total_rows - 1))
        self.current_col = max(0, min(self.current_col, self.max_cols - 1))
        if self.current_row < self.offset_row:
            self.offset_row = self.current_row
        if self.current_row >= self.offset_row + self.VIEW_ROWS:
            self.offset_row = max(0, self.current_row - self.VIEW_ROWS + 1)
        if self.current_col < self.offset_col:
            self.offset_col = self.current_col
        if self.current_col >= self.offset_col + self.VIEW_COLS:
            self.offset_col = max(0, self.current_col - self.VIEW_COLS + 1)
        self.reload_visible()

    def save(self) -> bool:
        out = self.filepath
        fmt = 'xlsx' if self.orig_format == 'xlsx' else 'csv'
        ok = self.fileman.finalize_and_write(out, format_hint=fmt, timeout=600)
        if ok:
            self.modifications.clear()
            self.deleted_cells.clear()
            self.modified = False
            self._recompute_metadata()
            self.reload_visible()
            self.status = "Saved successfully"
            return True
        else:
            self.status = "Save failed - restored from backup"
            return False

    def cleanup(self):
        try:
            self.fileman.cleanup()
        except Exception:
            pass

# ---------- search worker ----------
def _search_chunk(tempfile: str, start: int, end: int, term_l: str):
    res = []
    with open(tempfile, 'r', encoding='utf-8', newline='') as f:
        r = csv.reader(f)
        for i, row in enumerate(r):
            if i < start: continue
            if i >= end: break
            for c, val in enumerate(row):
                try:
                    if term_l in val.lower():
                        res.append((i,c))
                except Exception:
                    continue
    return res

# ---------- UI (curses) ----------
def draw_screen(stdscr, ed: Editor):
    stdscr.erase()
    h, w = stdscr.getmaxyx()
    title = f"{os.path.basename(ed.filepath)} {'*' if ed.modified else ''}"
    pos = f"Row {ed.current_row+1}/{ed.total_rows} Col {ed.current_col+1}/{ed.max_cols}"
    try:
        stdscr.addstr(0, 0, f"{title} | {pos}"[:w-1], curses.A_REVERSE)
    except curses.error:
        pass
    header_y = 1
    col_w = max(8, (w - 12) // ed.VIEW_COLS)
    for ci in range(ed.offset_col, ed.offset_col + ed.VIEW_COLS):
        x = 10 + (ci - ed.offset_col) * col_w
        if x >= w-1:
            break
        colname = chr(ord('A') + ci) if ci < 26 else f"C{ci+1}"
        try:
            stdscr.addstr(header_y, x, f"{colname:^{col_w-1}}", curses.A_BOLD)
        except curses.error:
            pass

    for r_off, row in enumerate(ed.visible_rows):
        y = header_y + 1 + r_off
        if y >= h - 3:
            break
        try:
            stdscr.addstr(y, 0, f"{ed.offset_row + r_off + 1:>6} | ")
        except curses.error:
            pass
        for ci in range(ed.offset_col, ed.offset_col + ed.VIEW_COLS):
            x = 10 + (ci - ed.offset_col) * col_w
            if x >= w-1:
                break
            cell = ''
            if ci < len(row):
                cell = row[ci]
            attr = 0
            if ed.selection_active and ed.sel_start and ed.sel_end:
                r1,c1 = ed.sel_start; r2,c2 = ed.sel_end
                top, bottom = min(r1,r2), max(r1,r2)
                left, right = min(c1,c2), max(c1,c2)
                rr = ed.offset_row + r_off
                if top <= rr <= bottom and left <= ci <= right:
                    attr |= curses.A_STANDOUT
            if (ed.offset_row + r_off) == ed.current_row and ci == ed.current_col:
                attr |= curses.A_REVERSE
            if (ed.offset_row + r_off, ci) in ed.modifications:
                attr |= curses.A_BOLD
            if (ed.offset_row + r_off, ci) in ed.deleted_cells:
                attr |= curses.A_DIM
            text = f"{cell[:col_w-2]:<{col_w-1}}"
            try:
                stdscr.addstr(y, x, text, attr)
            except curses.error:
                pass

    try:
        stdscr.hline(h-3, 0, '-', w)
    except curses.error:
        pass
    mode = ""
    if ed.selection_active:
        mode = "-- SELECT --"
    try:
        stdscr.addstr(h-2, 0, mode)
        stdscr.addstr(h-2, len(mode)+1, ed.status[:w - len(mode) - 2])
    except curses.error:
        pass
    help_line = "? help | s save | q quit"
    try:
        stdscr.addstr(h-1, 0, help_line[:w-1])
    except curses.error:
        pass
    stdscr.refresh()

def command_input(stdscr, prompt: str, prefill: str = "") -> str:
    h, w = stdscr.getmaxyx()
    y = h - 1
    try:
        stdscr.addstr(y, 0, " " * (w - 1))
        stdscr.addstr(y, 0, f"{prompt} {prefill}")
        stdscr.move(y, len(prompt) + 1 + len(prefill))
    except curses.error:
        pass
    curses.echo()
    curses.curs_set(1)
    buf = prefill
    try:
        while True:
            ch = stdscr.getch()
            if ch in (10, 13):
                break
            if ch in (27,):
                buf = ""
                break
            if ch in (8, 127):
                buf = buf[:-1]
            elif 0 <= ch < 256:
                buf += chr(ch)
            try:
                stdscr.addstr(y, 0, " " * (w - 1))
                stdscr.addstr(y, 0, f"{prompt} {buf}")
                stdscr.move(y, len(prompt) + 1 + len(buf))
                stdscr.refresh()
            except curses.error:
                pass
    finally:
        curses.noecho()
        curses.curs_set(0)
    return buf.strip()

def edit_cell_ui(stdscr, ed: Editor):
    ed.status = "Editing..."
    draw_screen(stdscr, ed)
    cur = ed.get_cell(ed.current_row, ed.current_col)
    new = command_input(stdscr, f"Edit [{ed.current_row+1},{ed.current_col+1}]:", cur)
    if new != cur and new != "":
        ed.set_cell(ed.current_row, ed.current_col, new)
        ed.status = f"Updated [{ed.current_row+1},{ed.current_col+1}]"
    else:
        ed.status = "Edit cancelled"

def auto_edit_cell(stdscr, ed: Editor, row: int, col: int, new_value: str):
    """Programmatic cell edit with UI consistency."""
    draw_screen(stdscr, ed)
    old = ed.get_cell(row, col)
    if new_value != old:
        ed.set_cell(row, col, new_value)
        ed.status = f"Updated [{row+1},{col+1}] (auto): '{str(new_value)[:40]}'"
    else:
        ed.status = f"No change [{row+1},{col+1}]"

def search_ui(stdscr, ed: Editor):
    ed.status = "Search..."
    draw_screen(stdscr, ed)
    term = command_input(stdscr, "/")
    if term:
        ed.search(term)
    else:
        ed.status = "Search cancelled"

import curses

def show_help_popup(stdscr):
    help_text = [
        " CSV Editor Help ",
        "",
        " ─ Navigation ─",
        "  ↑↓←→  Move cursor",
        "  PgUp / PgDn  Scroll screen",
        "  Home / End   Jump to start or end",
        "",
        " ─ Editing ─",
        "  e  Edit cell         d  Clear cell",
        "  r / R  Insert row    x  Delete row",
        "  C / V  Insert col    X  Delete col",
        "",
        " ─ Copy / Paste ─",
        "  ;  Start selection   Enter  Confirm selection",
        "  c  Copy              v  Paste",
        "",
        " ─ Search ─",
        "  /  Search text       n / N  Next / Prev match",
        "",
        " ─ Misc ─",
        "  g  Go to row         f  Full cell info",
        "  s  Save              :w, :q, :wq  Write/Quit",
        "  q  Quit (again to force)",
        "",
        "Press ? again to close help"
    ]

    h, w = stdscr.getmaxyx()
    box_h = min(len(help_text) + 2, h - 2)
    box_w = min(max(len(line) for line in help_text) + 4, w - 4)
    box_y = (h - box_h) // 2
    box_x = (w - box_w) // 2

    win = curses.newwin(box_h, box_w, box_y, box_x)
    win.keypad(True)
    top_line = 0

    while True:
        win.clear()
        win.box()
        visible_text = help_text[top_line: top_line + box_h - 2]
        for i, line in enumerate(visible_text, start=1):
            try:
                win.addstr(i, 2, line[:box_w - 4])
            except curses.error:
                pass
        win.refresh()

        key = win.getch()
        if key == ord('?'):
            break
        elif key == curses.KEY_UP:
            top_line = max(0, top_line - 1)
        elif key == curses.KEY_DOWN:
            if top_line + box_h - 2 < len(help_text):
                top_line += 1

def show_cell_popup(stdscr, ed):
    """Show full cell content in a centered popup, scrollable in both directions. Closes on F/f."""
    cell_value = ed.get_cell(ed.current_row, ed.current_col)
    cell_lines = cell_value.splitlines() or ["(empty)"]

    h, w = stdscr.getmaxyx()
    box_h = min(len(cell_lines) + 4, h - 2)
    max_line_len = max(len(line) for line in cell_lines)
    box_w = min(max_line_len + 6, w - 4)
    box_y = (h - box_h) // 2
    box_x = (w - box_w) // 2

    win = curses.newwin(box_h, box_w, box_y, box_x)
    win.keypad(True)
    top_line = 0
    left_offset = 0

    while True:
        win.clear()
        win.box()

        title = f" Cell [{ed.current_row+1},{ed.current_col+1}] "
        try:
            win.addstr(0, max(2, (box_w - len(title)) // 2), title, curses.A_REVERSE)
        except curses.error:
            pass

        # Draw visible portion of lines
        visible_lines = cell_lines[top_line: top_line + box_h - 3]
        for i, line in enumerate(visible_lines, start=1):
            segment = line[left_offset:left_offset + (box_w - 4)]
            try:
                win.addstr(i, 2, segment)
            except curses.error:
                pass

        # Optional scroll indicator
        if left_offset > 0:
            win.addstr(box_h - 2, 2, "←", curses.A_DIM)
        if left_offset + (box_w - 4) < max_line_len:
            win.addstr(box_h - 2, box_w - 3, "→", curses.A_DIM)

        win.refresh()

        key = win.getch()
        if key in (ord('f'), ord('F')):
            break
        elif key == curses.KEY_UP:
            top_line = max(0, top_line - 1)
        elif key == curses.KEY_DOWN:
            if top_line + box_h - 3 < len(cell_lines):
                top_line += 1
        elif key == curses.KEY_LEFT:
            left_offset = max(0, left_offset - 4)
        elif key == curses.KEY_RIGHT:
            if left_offset + (box_w - 4) < max_line_len:
                left_offset += 4

    del win
    stdscr.touchwin()
    stdscr.refresh()
    draw_screen(stdscr, ed)




# ---------- Main loop ----------
def main(stdscr):
    if len(sys.argv) < 2:
        print("Usage: python3 csv_edit.py <file.csv|file.xlsx>")
        return
    curses.curs_set(0)
    stdscr.timeout(100)
    path = sys.argv[1]
    if not os.path.exists(path):
        open(path, 'w', encoding='utf-8').close()

    filename, ext = os.path.splitext(path)

    ed = Editor(path)
    ed.status = "Ready - ? for help" if ext == ".csv" else "WARNING: SOME DATA MAY BE LOST DUE TO CSV LIMITATIONS"

    try:
        while True:
            draw_screen(stdscr, ed)
            key = stdscr.getch()
            if key == -1:
                continue

            if key == ord('q'):
                if ed.modified:
                    ed.status = "Unsaved changes! Use :w to save, or q again to quit (will lose edits)"
                    second = stdscr.getch()
                    if second == ord('q'):
                        break
                else:
                    break
            elif key == ord(':'):
                cmd = command_input(stdscr, ":")
                if cmd in ('q', 'quit'):
                    if ed.modified:
                        ed.status = "Unsaved changes! Use :w to save"
                    else:
                        break
                elif cmd in ('w','write','save'):
                    ed.save()
                elif cmd in ('wq','x'):
                    ed.save()
                    break
                elif cmd == 'q!':
                    break
                else:
                    ed.status = f"Unknown command: {cmd}"

            elif key == curses.KEY_UP:
                ed.current_row = max(0, ed.current_row - 1)
                if ed.current_row < ed.offset_row:
                    ed.offset_row = ed.current_row
                    ed.reload_visible()
                if ed.selection_active:
                    ed.expand_selection_to_cursor()
            elif key == curses.KEY_DOWN:
                ed.current_row = min(ed.total_rows - 1, ed.current_row + 1)
                if ed.current_row >= ed.offset_row + ed.VIEW_ROWS:
                    ed.offset_row = ed.current_row - ed.VIEW_ROWS + 1
                    ed.reload_visible()
                if ed.selection_active:
                    ed.expand_selection_to_cursor()
            elif key == curses.KEY_LEFT:
                ed.current_col = max(0, ed.current_col - 1)
                if ed.current_col < ed.offset_col:
                    ed.offset_col = ed.current_col
                    ed.reload_visible()
                if ed.selection_active:
                    ed.expand_selection_to_cursor()
            elif key == curses.KEY_RIGHT:
                ed.current_col = min(ed.max_cols - 1, ed.current_col + 1)
                if ed.current_col >= ed.offset_col + ed.VIEW_COLS:
                    ed.offset_col = ed.current_col - ed.VIEW_COLS + 1
                    ed.reload_visible()
                if ed.selection_active:
                    ed.expand_selection_to_cursor()

            elif key == curses.KEY_PPAGE:
                ed.current_row = max(0, ed.current_row - ed.VIEW_ROWS)
                ed.offset_row = max(0, ed.offset_row - ed.VIEW_ROWS)
                ed.reload_visible()
            elif key == curses.KEY_NPAGE:
                ed.current_row = min(ed.total_rows - 1, ed.current_row + ed.VIEW_ROWS)
                ed.offset_row = min(max(0, ed.total_rows - ed.VIEW_ROWS), ed.offset_row + ed.VIEW_ROWS)
                ed.reload_visible()
            elif key == curses.KEY_HOME:
                ed.current_col = 0
                ed.offset_col = 0
                ed.reload_visible()
            elif key == curses.KEY_END:
                ed.current_col = max(0, ed.max_cols - 1)
                ed.offset_col = max(0, ed.max_cols - ed.VIEW_COLS)
                ed.reload_visible()

            elif key == ord(';'):
                ed.start_selection()
            elif key in (10,13):  # Enter
                if ed.selection_active:
                    ed.confirm_selection_copy()
                    ed.reload_visible()
                else:
                    edit_cell_ui(stdscr, ed)
            elif key == 27:  # Esc
                if ed.selection_active:
                    ed.cancel_selection()

            elif key == ord('e'):
                edit_cell_ui(stdscr, ed)
            elif key == ord('/'):
                search_ui(stdscr, ed)
            elif key == ord('n'):
                if ed.search_next():
                    ed.ensure_visible()
                    ed.status = f"Search match at [{ed.current_row+1},{ed.current_col+1}]"
                else:
                    ed.status = "No further matches"
            elif key == ord('N'):
                if ed.search_prev():
                    ed.ensure_visible()
                    ed.status = f"Search match at [{ed.current_row+1},{ed.current_col+1}]"
                else:
                    ed.status = "No previous matches"
            elif key == ord('s'):
                ed.save()
            elif key == ord('d'):
                old = ed.get_cell(ed.current_row, ed.current_col)
                ed.clear_cell(ed.current_row, ed.current_col)
                ed.reload_visible()
                ed.status = f"Deleted [{ed.current_row+1},{ed.current_col+1}] = '{old[:20]}'"
            elif key == ord('c'):
                ed.copy_single_or_selection()
                ed.status = "Copied selection"
            elif key == ord('v'):
                pasted = ed.paste_overwrite()
                if pasted:
                    for (r, c, val) in pasted:
                        auto_edit_cell(stdscr, ed, r, c, val)
                    ed.status = f"Pasted {len(pasted)} cell(s)"
                else:
                    ed.status = "Nothing to paste"
            elif key == ord('r'):
                ed.insert_row_below()
                ed.reload_visible()
            elif key == ord('R'):
                ed.insert_row_above()
                ed.reload_visible()
            elif key == ord('x'):
                ed.delete_current_row()
                ed.reload_visible()
            elif key == ord('C'):
                ed.insert_col_right()
                ed.reload_visible()
            elif key == ord('V'):
                ed.insert_col_left()
                ed.reload_visible()
            elif key == ord('F'):
                show_cell_popup(stdscr, ed)
            elif key == ord('X'):
                ed.delete_current_col()
                ed.reload_visible()
            elif key == ord('g'):  # 'g' for goto
                row_str = command_input(stdscr, "Go to row number:")
                try:
                    target = int(row_str)
                    ed.goto_row(target)
                except ValueError:
                    ed.status = "Invalid row number"
            elif key == ord('f'):
                full_value = ed.get_cell(ed.current_row, ed.current_col)
                ed.status = f"Full cell [{ed.current_row+1},{ed.current_col+1}]: {full_value}"
            elif key == ord('?'):
                show_help_popup(stdscr)
                ed.status = "Ready - ? for help"
            else:
                ed.status = "Ready - ? for help"

    finally:
        ed.cleanup()


if __name__ == '__main__':
    curses.wrapper(main)
